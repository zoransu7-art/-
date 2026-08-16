'''
ctrip_spider2:携程评论爬虫第3个进程
根据ctrip_spider1给的详情页网址获取景点序号poiId
然后根据poiId获取对应游客评论并保存到excel
新增功能：
1. 爬取完不足300页的景点存入{城市}_未达已爬取.txt
2. 爬取前检查该文件，跳过已记录的景点
'''
import openpyxl
import time
import requests
import json
import random
import csv
import keyboard  # 用于键盘控制（暂停/中断）
from fake_useragent import UserAgent
import uuid
from bs4 import BeautifulSoup
import re
from datetime import datetime  # 用于记录评论时间
import pandas as pd
import os
from openpyxl import load_workbook
import traceback

# ===================== 核心优化：分域会话管理 =====================
# 为不同域名创建专用会话，初始化为None，首次请求时创建
POI_SESSION = None  # 专属you.ctrip.com（poiId提取）的会话
COMMENT_SESSION = None  # 专属m.ctrip.com（评论爬取）的会话

def create_poi_session():
    """创建/重建poiId提取的专用会话（you.ctrip.com）"""
    global POI_SESSION
    POI_SESSION = requests.Session()
    # 为poi会话统一设置基础请求头，避免重复代码
    POI_SESSION.headers.update({
        'authority': 'you.ctrip.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-encoding': 'gzip, deflate, br, zstd',
        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'cache-control': 'max-age=0',
        'priority': 'u=0, i',
        'sec-ch-ua': '"Google Chrome";v="141", "Not?A_Brand";v="8", "Chromium";v="141"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
    })
    return POI_SESSION

def create_comment_session():
    """创建/重建评论爬取的专用会话（m.ctrip.com）"""
    global COMMENT_SESSION
    COMMENT_SESSION = requests.Session()
    # 为评论会话统一设置基础请求头
    COMMENT_SESSION.headers.update({
        'authority': 'm.ctrip.com',
        'accept': '*/*',
        'accept-encoding': 'gzip, deflate, br, zstd',
        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'content-type': 'application/json; charset=utf-8',
        'cookieorigin': 'https://you.ctrip.com',
        'origin': 'https://you.ctrip.com',
        'priority': 'u=1, i',
        'sec-ch-ua': '"Google Chrome";v="141", "Not?A_Brand";v="8", "Chromium";v="141"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-site',
    })
    return COMMENT_SESSION

# ===================== 新增函数：未达300页景点管理 =====================
def get_skipped_spots(city):
    """
    读取{城市}_未达已爬取.txt文件，获取需要跳过的景点列表
    :param city: 城市名称
    :return: 包含跳过景点名称的集合
    """
    skipped_file = f"{city}_未达已爬取.txt"
    skipped_spots = set()

    if os.path.exists(skipped_file):
        try:
            with open(skipped_file, 'r', encoding='utf-8') as f:
                for line in f:
                    spot_name = line.strip()
                    if spot_name:
                        skipped_spots.add(spot_name)
            print(f"从 {skipped_file} 读取到 {len(skipped_spots)} 个需要跳过的景点")
        except Exception as e:
            print(f"读取跳过景点文件失败: {e}")
    else:
        print(f"未找到 {skipped_file} 文件，将创建新文件（如有需要）")

    return skipped_spots

def save_spot_to_skipped_list(city, spot_name, total_pages):
    """
    将爬取完但总页数不足300页的景点保存到{城市}_未达已爬取.txt
    :param city: 城市名称
    :param spot_name: 景点名称
    :param total_pages: 实际爬取的总页数
    """
    skipped_file = f"{city}_未达已爬取.txt"
    # 格式化写入内容，包含页数信息便于后续查看
    content = f"{spot_name}\t总页数：{total_pages}\t记录时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"

    try:
        with open(skipped_file, 'a', encoding='utf-8') as f:
            f.write(content)
        print(f"已将景点 '{spot_name}' (总页数{total_pages}) 记录到 {skipped_file}")
    except Exception as e:
        print(f"保存跳过景点失败: {e}")

# ===================== 原有函数优化 =====================
def read_spots_from_csv(city):
    """
    从 {城市}_携程景点.csv 读取景点名称和详情页URL
    :param city: 城市名称
    :return: 包含景点名称和URL的列表 [(spot_name, detail_url), ...]
    """
    csv_file = f"{city}_携程景点_重合.csv"
    spots_data = []

    try:
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                spot_name = row.get("景点", "").strip()
                detail_url = row.get("详情页URL", "").strip()

                # 只有当URL不为空时才处理
                if spot_name and detail_url:
                    spots_data.append((spot_name, detail_url))
                elif spot_name and not detail_url:
                    print(f"跳过景点 '{spot_name}'：无详情页URL")

        print(f"从 {csv_file} 读取到 {len(spots_data)} 个有效景点")
        return spots_data

    except FileNotFoundError:
        print(f"错误: 未找到文件 {csv_file}，请先运行 ctrip_spider0.py 和 ctrip_spider1.py")
        return []
    except Exception as e:
        print(f"读取 {csv_file} 时出错: {e}")
        return []

def get_poiId(url, max_retry=3, retry_interval=10):
    """
    从景点详情页HTML中提取poiId（优化：分域会话+重试重建连接）
    :param url: 景点详情页URL
    :return: poiId（字符串格式，如"13412802"）
    """
    # 首次请求/重试时，创建/重建poi会话（换连接）
    poi_session = create_poi_session()

    for retry_count in range(1, max_retry + 1):
        # 每次重试都动态更换UA和Cookie中的唯一标识（反爬关键）
        random_ua = UserAgent().chrome
        random_guid = str(uuid.uuid4())

        # 动态构造Cookie
        cookies = {
            'GUID': random_guid[:20],
            'nfes_isSupportWebP': '1',
            'UBT_VID': f'{int(time.time()*1000)}.{random_guid[:8]}',
            'MKT_CKID': f'{int(time.time()*1000)}.{random_guid[:4]}.{random_guid[4:8]}',
            '_RGUID': random_guid,
            '_bfaStatusPVSend': '1',
            'Hm_lvt_e4211314613fcf074540918eb10eeecb': f'{int(time.time()-86400)},{int(time.time()-3600)},{int(time.time())}',
            'HMACCOUNT': ''.join(random.choice('0123456789ABCDEF') for _ in range(16)),
            '_bfaStatus': 'success',
        }

        try:
            print(f"\n--- 第{retry_count}/{max_retry}次尝试提取poiId ---")
            # 更新当前会话的UA
            poi_session.headers['user-agent'] = random_ua

            # 使用poi专用会话发请求，成功则复用连接，失败则下次重试重建
            response = poi_session.get(
                url,
                cookies=cookies,
                timeout=30,
                allow_redirects=True  # 处理携程的重定向反爬
            )

            if response.status_code == 200:
                soup = BeautifulSoup(response.text, "html.parser")
                script = soup.find('script', id='__NEXT_DATA__')

                if script:
                    next_data = json.loads(script.string)
                    # 尝试多种方式提取poiId
                    poi_id = (
                            next_data.get('poiId') or
                            next_data.get('props', {}).get('pageProps', {}).get('poiId') or
                            next_data.get('props', {}).get('pageProps', {}).get('initialState', {}).get('baseData', {}).get('poiId') or
                            next_data.get('props', {}).get('pageProps', {}).get('initialState', {}).get('poiDetail', {}).get('poiId')
                    )

                    if poi_id:
                        print(f"=== poiId提取成功（第{retry_count}次尝试）===")
                        print(f"poiId: {poi_id}")
                        return str(poi_id)
                    else:
                        print("poiId未找到，数据结构可能更新")
                        print(f"1. next_data.get('poiId'): {next_data.get('poiId')}")
                        print(f"2. pageProps.poiId: {next_data.get('props', {}).get('pageProps', {}).get('poiId')}")
                else:
                    print("未找到 __NEXT_DATA__ 脚本标签（反爬拦截可能）")

            else:
                print(f"HTTP状态码异常: {response.status_code}（反爬拦截可能）")

        except Exception as e:
            print(f"请求详情页 {url} 失败：{e}")
            # 异常时主动重建会话（确保下次重试用新连接）
            create_poi_session()

        # 指数退避重试，且重试前重建会话（核心：换连接）
        if retry_count < max_retry:
            backoff = retry_interval * (2 ** (retry_count - 1))  # 指数退避
            print(f"等待{backoff}秒后重试（已重建新连接）...")
            time.sleep(backoff)
            # 重试前强制重建会话，换全新TCP连接
            create_poi_session()

    print(f"=== 已达到最大重试次数（{max_retry}次），poiId提取失败 ===")
    return None

# 补充缺失的 clean_text 函数
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\n\r\t]', ' ', str(text)).strip()

def get_comments(poiid, detail_url, file_path, scenery_name, stop_flag, city):
    """
    评论采集（优化：分域会话+批量保存+动态延迟+记录不足300页景点）
    1. 起始页 = 已爬页数 + 1；
    2. 终止条件：爬满300页 / 某页无评论 / ESC中断；
    3. 新增：记录实际爬取总页数，不足300页则保存到跳过列表
    """
    url = "https://m.ctrip.com/restapi/soa2/13444/json/getCommentCollapseList"
    batch_comments = []  # 存储批量评论（每10条保存一次）
    batch_size = 10      # 批量保存阈值
    is_running = True
    is_paused = False
    total_crawled_pages = 0  # 新增：记录实际爬取的总页数

    # 生成固定GUID（避免每次请求换GUID，降低反爬风险）
    guid = ''.join(str(random.randint(0, 9)) for _ in range(20))

    # 键盘事件监听：空格=暂停/继续，ESC=中断
    def on_space_press():
        nonlocal is_paused
        is_paused = not is_paused
        print("\n 已暂停采集，按空格键继续..." if is_paused else "\n 已继续采集...")

    keyboard.on_press_key("space", lambda _: on_space_press())
    print("采集控制说明：按空格键暂停/继续，按ESC键中断采集")

    # 获取起始页
    existing_count = 0
    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path, read_only=True)
            if scenery_name in book.sheetnames:
                existing_count = max(0, book[scenery_name].max_row - 1)
            book.close()
        except Exception as e:
            print(f"读取已爬页数出错: {e}")
            existing_count = 0

    if existing_count > 0:
        if existing_count % 10 != 0:  # 余数不为0
            start_page = (existing_count // 10) + 2
        else:  # 余数为0，正好是整页
            start_page = (existing_count // 10) + 1
    else:
        start_page = 1
    # start_page = (existing_count // 10) + 1 if existing_count > 0 else 1

    try:
        for page in range(start_page, 301):
            # 检查全局停止标志
            if stop_flag.get('stop', False):
                break

            # 暂停逻辑：暂停时也持续检查停止标志
            while is_paused and not stop_flag.get('stop', False):
                time.sleep(0.5)
            if stop_flag.get('stop', False):
                break

            # 构造请求参数
            json_data = {
                'head': {
                    'auth': '',
                    'cid': guid,
                    'ctok': '',
                    'cver': '1.0',
                    'extension': [],
                    'lang': '01',
                    'sid': '8888',
                    'syscode': '09',
                    'xsid': '',
                },
                'arg': {
                    'channelType': 2,
                    'collapseType': 0,
                    'commentTagId': 0,
                    'pageIndex': page,
                    'pageSize': 10,
                    'poiId': poiid,
                    'sourceType': 1,
                    'sortType': 3,
                    'starType': 0,
                },
            }

            # 动态生成参数
            timestamp = int(time.time() * 1000)
            x_trace_id = f"{guid}-{timestamp}-{random.randint(1000000, 9999999)}"

            params = {
                '_fxpcqlniredt': guid,
                'x-traceID': x_trace_id
            }

            # 构造Cookie
            cookies = {
                'GUID': guid,
                'nfes_isSupportWebP': '1',
                '_RGUID': str(uuid.uuid4()),
                '_bfaStatus': 'success',
            }

            try:
                # 首次请求/失败后重建评论专用会话
                comment_session = create_comment_session()
                # 动态设置UA和Referer
                comment_session.headers['user-agent'] = UserAgent().chrome
                comment_session.headers['referer'] = detail_url

                # 构造请求体（避免中文转义）
                json_str = json.dumps(json_data, ensure_ascii=False)
                json_bytes = json_str.encode('utf-8')

                # 使用评论专用会话发请求（复用连接）
                response = comment_session.post(
                    url,
                    params=params,
                    cookies=cookies,
                    data=json_bytes,
                    timeout=30
                )
                response.encoding = 'utf-8'  # 解决中文编码问题
                response.raise_for_status()  # 若状态码非200，直接抛出异常

                # 解析响应
                data = response.json()

                # 检查API响应状态
                if data.get("ResponseStatus", {}).get("Ack") == "Success":
                    # 提取items列表
                    items = data.get("result", {}).get("items", [])
                    if not items:
                        print(f"第{page}页无评论数据，可能已采集完")
                        total_crawled_pages = page - 1  # 记录实际爬取页数
                        break

                    for item in items:
                        user_info = item.get("userInfo", {}) or {}  # 处理userInfo为null的情况

                        # 提取评论发布时间
                        publishTypeTag = item.get("publishTypeTag", None)
                        publishTime = re.search(r"\d{4}-\d{2}-\d{2}", publishTypeTag).group(0) if publishTypeTag else None

                        # 获取各类评分信息
                        scores = item.get("scores", [])
                        scenic_score = next((s['score'] for s in scores if s['name'] == '景色'), None)
                        fun_score = next((s['score'] for s in scores if s['name'] == '趣味'), None)
                        cost_performance_score = next((s['score'] for s in scores if s['name'] == '性价比'), None)

                        review_info = {
                            "用户昵称": user_info.get("userNick", "未知"),
                            "用户评分": item.get("score", {}),
                            "评论时间": publishTime,
                            "IP属地": item.get("ipLocatedName", "未知"),
                            "游客类型": item.get("touristTypeDisplay", "未知"),
                            "景色评分": scenic_score,
                            "趣味评分": fun_score,
                            "性价比评分": cost_performance_score,
                            "评论内容": clean_text(item.get("content", "")),
                        }
                        batch_comments.append(review_info)

                    total_crawled_pages = page  # 更新已爬页数
                    print(f"第{page}页采集完成（共{len(items)}条）")

                    # 批量保存：达到batch_size或最后一页时保存
                    if len(batch_comments) >= batch_size or page == 300 or not items:
                        save_comments_batch_to_excel(batch_comments, file_path, scenery_name)
                        batch_comments = []  # 清空批量列表

                else:
                    error_msg = data.get("ResponseStatus", {}).get("Errors", [{}])[0].get("Message", "未知错误")
                    print(f"第{page}页接口返回失败：{clean_text(error_msg)}")

            except Exception as e:
                print(f"\n第{page}页采集出错：{type(e).__name__} - {str(e)}")
                # 评论请求失败：重建会话（换连接）+ 延迟重试
                create_comment_session()
                time.sleep(random.uniform(2, 5))
                continue

            # 动态延迟：根据请求结果调整，避免固定长延迟
            time.sleep(random.uniform(0.5, 1.5))

        # 处理爬满300页的情况
        if total_crawled_pages == 0:
            total_crawled_pages = start_page - 1  # 未爬取任何新页面
        elif total_crawled_pages < 300 and not stop_flag.get('stop', False):
            # 只有在非手动停止且页数不足300时才记录
            save_spot_to_skipped_list(city, scenery_name, total_crawled_pages)

    finally:
        # 保存剩余的评论
        if batch_comments:
            save_comments_batch_to_excel(batch_comments, file_path, scenery_name)
        # 注销键盘监听
        keyboard.unhook_key("space")
        print(f"\n景点 '{scenery_name}' 采集结束！实际爬取总页数：{total_crawled_pages}")

# ===================== Excel操作优化 =====================
def init_excel(file_path, scenery_name):
    """初始化Excel文件：优化创建逻辑"""
    headers = ["景点名称", "用户昵称", "用户评分", "评论时间", "IP属地",
               "游客类型", "景色评分", "趣味评分", "性价比评分", "评论内容"]

    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
            if scenery_name not in book.sheetnames:
                new_sheet = book.create_sheet(scenery_name)
                for col_idx, header in enumerate(headers, 1):
                    new_sheet.cell(row=1, column=col_idx, value=header)
                book.save(file_path)
                book.close()
                print(f"已为景点 '{scenery_name}' 创建新的sheet")
            else:
                book.close()
            return True
        except Exception as e:
            print(f"处理现有Excel文件时出错: {e}")
            return False
    else:
        try:
            book = openpyxl.Workbook()
            default_sheet = book.active
            book.remove(default_sheet)
            sheet = book.create_sheet(scenery_name)
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_idx, value=header)
            book.save(file_path)
            book.close()
            print(f"已创建Excel文件: {file_path}")
            return True
        except Exception as e:
            print(f"创建Excel文件时出错: {e}")
            return False

def save_comments_batch_to_excel(comments, file_path, scenery_name):
    """批量保存评论到Excel（核心优化点）"""
    if not comments:
        return

    full_columns = ["景点名称", "用户昵称", "用户评分", "评论时间", "IP属地",
                    "游客类型", "景色评分", "趣味评分", "性价比评分", "评论内容"]

    try:
        book = load_workbook(file_path)
        sheet = book[scenery_name]
        start_row = sheet.max_row + 1

        # 批量写入数据
        for row_idx, comment in enumerate(comments, start_row):
            comment_with_scenery = comment.copy()
            comment_with_scenery["景点名称"] = scenery_name

            for col_idx, col_name in enumerate(full_columns, 1):
                value = comment_with_scenery.get(col_name, "")
                sheet.cell(row=row_idx, column=col_idx, value=value)

        book.save(file_path)
        book.close()
        # print(f"✓ 批量保存 {len(comments)} 条评论到Excel")

    except Exception as e:
        print(f"批量保存评论到Excel出错: {e}")
        traceback.print_exc()

# ===================== 辅助函数 =====================
def find_available_cities():
    """查找所有可用的城市文件"""
    import glob
    city_files = glob.glob("*_携程景点_重合.csv")
    cities = []
    for file in city_files:
        city_name = file.replace("_携程景点_重合.csv", "")
        cities.append((city_name, file))
    return cities

def get_remaining_spots(csv_file, excel_file, skipped_spots):
    """
    获取需要爬取的剩余景点（排除已跳过的景点）
    :param csv_file: CSV文件路径
    :param excel_file: Excel文件路径
    :param skipped_spots: 需要跳过的景点集合
    :return: 剩余景点列表
    """
    all_spots = []
    try:
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                spot_name = row.get("景点", "").strip()
                detail_url = row.get("详情页URL", "").strip()
                if spot_name and detail_url:
                    all_spots.append((spot_name, detail_url))
        print(f"从CSV读取到 {len(all_spots)} 个有效景点")
    except Exception as e:
        print(f"读取CSV文件失败: {e}")
        return all_spots

    # 第一步：排除已记录在跳过列表的景点
    all_spots = [(name, url) for name, url in all_spots if name not in skipped_spots]
    print(f"排除跳过列表中的景点后剩余: {len(all_spots)} 个")

    if not os.path.exists(excel_file):
        print("Excel文件不存在，将爬取所有剩余景点")
        return all_spots

    # 第二步：排除Excel中已有完整数据的景点
    crawled_spots_with_data = set()
    try:
        book = load_workbook(excel_file, read_only=True)
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            comment_count = max(0, sheet.max_row - 1)
            if comment_count >= 2900:
                crawled_spots_with_data.add(sheet_name)
        book.close()
        print(f"Excel中已有 {len(crawled_spots_with_data)} 个景点有完整数据")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return all_spots

    remaining_spots = [
        (name, url) for name, url in all_spots
        if name not in crawled_spots_with_data
    ]
    print(f"最终剩余需要爬取的景点: {len(remaining_spots)} 个")
    return remaining_spots

# ===================== 主函数 =====================
def mainc2():
    """主函数：优化整体流程+分域会话+资源释放+新增跳过景点逻辑"""
    # 1. 自动检测可用的城市文件
    available_cities = find_available_cities()

    if not available_cities:
        print("未找到任何城市景点文件，请先运行 ctrip_spider0.py 和 ctrip_spider1.py")
        return

    # 2. 选择城市
    if len(available_cities) == 1:
        city, csv_file = available_cities[0]
        print(f"自动检测到城市: {city}")
    else:
        print("检测到多个城市文件，请选择:")
        for i, (city_name, file_path) in enumerate(available_cities, 1):
            print(f"{i}. {city_name}")

        try:
            choice = int(input("请输入序号: ")) - 1
            if 0 <= choice < len(available_cities):
                city, csv_file = available_cities[choice]
            else:
                print("选择无效，请手动输入城市名称")
                city = input("请输入城市名称: ").strip()
                csv_file = f"{city}_携程景点.csv"
        except ValueError:
            print("输入无效，请手动输入城市名称")
            city = input("请输入城市名称: ").strip()
            csv_file = f"{city}_携程景点.csv"

    # 3. 新增：读取需要跳过的景点列表
    skipped_spots = get_skipped_spots(city)

    # 4. 配置保存路径
    platform = "携程"
    excel_file_path = f"{city}_{platform}评论.xlsx"

    # 5. 获取需要爬取的剩余景点（传入跳过列表）
    remaining_spots = get_remaining_spots(csv_file, excel_file_path, skipped_spots)

    if not remaining_spots:
        print("所有景点均已爬取完成！")
        # 显示统计信息
        try:
            book = load_workbook(excel_file_path, read_only=True)
            total_comments = 0
            for sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                comment_count = sheet.max_row - 1
                if comment_count > 0:
                    print(f"{sheet_name}: {comment_count} 条评论")
                    total_comments += comment_count
            print(f"\n总计: {len(book.sheetnames)} 个景点, {total_comments} 条评论")
            book.close()
        except:
            print("无法读取统计信息")
        return

    print(f"\n开始处理 {len(remaining_spots)} 个剩余景点...")

    # 6. 全局停止标志
    stop_flag = {'stop': False}

    def on_esc_press():
        stop_flag['stop'] = True
        print("\nESC键按下，将停止所有采集（当前景点评论页完成后生效）...")

    # 注册全局ESC监听
    keyboard.on_press_key("esc", lambda _: on_esc_press())
    print("全局控制：按ESC键可中断所有景点的采集")

    # 7. 遍历每个剩余景点进行评论采集
    try:
        for i, (spot_name, detail_url) in enumerate(remaining_spots, 1):
            # 检查全局停止标志
            if stop_flag.get('stop', False):
                print("\n已停止所有采集")
                break

            print(f"\n{'=' * 60}")
            print(f"[{i}/{len(remaining_spots)}] 正在处理景点: {spot_name}")

            # 提取poiId
            print("正在提取poiId...")
            poiid = get_poiId(detail_url, max_retry=3, retry_interval=10)

            if not poiid:
                print(f"✗ 未提取到poiId，跳过景点 '{spot_name}'")
                # 记录失败景点
                with open(f"failed_spots_{city}.txt", "a", encoding="utf-8") as f:
                    f.write(f"{spot_name}\n")
                continue

            # 初始化Excel
            if not init_excel(excel_file_path, spot_name):
                print(f"✗ 初始化Excel失败，跳过景点 '{spot_name}'")
                continue

            # 开始采集评论（新增传入city参数）
            print(f"开始采集景点 '{spot_name}' 的评论...")
            get_comments(
                poiid=poiid,
                detail_url=detail_url,
                file_path=excel_file_path,
                scenery_name=spot_name,
                stop_flag=stop_flag,
                city=city  # 新增传入城市名，用于保存跳过列表
            )

            # 检查是否被ESC中断
            if stop_flag['stop']:
                break

            # 景点间动态延迟
            if i < len(remaining_spots):
                delay = random.uniform(1, 3)
                print(f"等待 {delay:.1f} 秒后处理下一个景点...")
                time.sleep(delay)
    finally:
        # 释放资源
        keyboard.unhook_key("esc")
        keyboard.unhook_all()
        # 关闭所有会话
        global POI_SESSION, COMMENT_SESSION
        if POI_SESSION:
            POI_SESSION.close()
        if COMMENT_SESSION:
            COMMENT_SESSION.close()

    print(f"\n{'=' * 60}")
    print(f"采集完成！评论数据已保存到: {excel_file_path}")

'''
遇到的问题:
1.保存为csv时有些emoj无法被识别，所以修改为excel了
2.pandas 的 FutureWarning（未来警告），一个版本兼容性警告，意思是：
当前版本：pandas 在合并空数据列时会自动排除它们
貌似是因为有空值，但是本身有些数据列就是空的
解决方法:忽略警告
3.有时无法获取poiid，但没有报错所以不知道怎么回事，但猜测是爬取间断太短被反爬机制制止了
(9/18的保存度，最大间隔15min)
已解决:新增间隔机制，但问题好像是因为提取poiid的代码有问题，通过检索返回的数据，发现结构和网页观察到的不同，已修改，未再出错

4.没有设置防中断机制，如果发生问题，重新跑代码将从头开始
已解决:需新增检测机制，对比景点文件跳过已经爬取成功的景点评论，爬取未成功的

5.爬取效率太慢
解决方案:
1)使用TCP连接，poiId 提取和评论爬取的域名不同、且反爬重试需要断开重连，直接用一个全局会话将存在适配性问题，
分域创建独立会话池 + 重试时重建会话
为you.ctrip.com（poiId）和m.ctrip.com（评论）各创建一个专用会话对象，各自维护独立的连接池，实现同域请求的连接复用，跨域互不影响；
2)将1条评论一存改为10条(1页)，load_workbook后没有及时关闭工作簿（虽然save后会自动关，但频繁加载仍会残留内存占用）
'''